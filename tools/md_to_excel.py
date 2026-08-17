# -*- coding: utf-8 -*-
"""
将 DMS 测试用例 Markdown 文档转换为 Excel 文件
每个章节（## 第X章 / ## 附录X）作为一个 sheet
"""
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = Path(r"d:\Workspace\TRAE\DMS\docs\10_测试用例\DMS完整测试场景与测试案例_v3.11.0.md")
DST = Path(r"d:\Workspace\TRAE\DMS\docs\10_测试用例\DMS完整测试场景与测试案例_v3.11.0.xlsx")

# 样式
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
SECTION_FONT = Font(name="微软雅黑", size=12, bold=True, color="1F4E78")
SUBSECTION_FONT = Font(name="微软雅黑", size=11, bold=True, color="2E75B6")
NORMAL_FONT = Font(name="微软雅黑", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_table(lines):
    """解析 markdown 表格行，返回二维数组（含表头）"""
    rows = []
    for ln in lines:
        ln = ln.strip()
        if not ln.startswith("|"):
            continue
        # 跳过分隔行
        if re.match(r"^\|[\s\-:|]+\|$", ln):
            continue
        cells = [c.strip() for c in ln.strip("|").split("|")]
        rows.append(cells)
    return rows


def sanitize_sheet_name(name):
    """Excel sheet 名限制：≤31字符，不能包含 : \ / ? * [ ] """
    name = re.sub(r"[:\\/?*\[\]]", "-", name)
    if len(name) > 31:
        name = name[:31]
    return name


def write_table(ws, start_row, table_rows, table_num):
    """将一个表格写入 worksheet，返回下一个空行"""
    # 表格序号
    ws.cell(row=start_row, column=1, value=f"[表格 {table_num}]").font = SUBSECTION_FONT
    start_row += 1

    if not table_rows:
        return start_row

    # 表头
    header = table_rows[0]
    for col_idx, val in enumerate(header, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    start_row += 1

    # 数据行
    for row in table_rows[1:]:
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_WRAP
            cell.border = BORDER
        start_row += 1

    return start_row + 1  # 空一行


def write_section_info(ws, row, title, level):
    """写入章节标题"""
    cell = ws.cell(row=row, column=1, value=title)
    if level == 1:
        cell.font = Font(name="微软雅黑", size=14, bold=True, color="C00000")
    elif level == 2:
        cell.font = SECTION_FONT
    else:
        cell.font = SUBSECTION_FONT
    return row + 2


def main():
    print(f"读取源文件: {SRC}")
    content = SRC.read_text(encoding="utf-8")
    lines = content.split("\n")
    print(f"总行数: {len(lines)}")

    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    # 按 ## 分章节
    chapters = []
    current_chapter = {"title": "文档头部", "lines": []}
    for ln in lines:
        if ln.startswith("## "):
            chapters.append(current_chapter)
            current_chapter = {"title": ln[3:].strip(), "lines": []}
        else:
            current_chapter["lines"].append(ln)
    chapters.append(current_chapter)

    print(f"共 {len(chapters)} 个章节")

    table_total = 0
    for ch_idx, chapter in enumerate(chapters):
        title = chapter["title"]
        ch_lines = chapter["lines"]
        sheet_name = sanitize_sheet_name(title) if title != "文档头部" else "文档头部"
        # 如果同名已存在，加序号
        base_name = sheet_name
        suffix = 1
        while sheet_name in [s.title for s in wb.worksheets]:
            suffix += 1
            sheet_name = sanitize_sheet_name(f"{base_name}_{suffix}")

        ws = wb.create_sheet(title=sheet_name)
        row = 1
        # 章节标题
        row = write_section_info(ws, row, title, 1)

        # 遍历行，识别 ### 子章节 和 表格
        i = 0
        table_num = 0
        while i < len(ch_lines):
            ln = ch_lines[i]
            stripped = ln.strip()

            # 子章节标题
            if stripped.startswith("### "):
                row = write_section_info(ws, row, stripped[4:].strip(), 2)
                i += 1
                continue
            if stripped.startswith("#### "):
                row = write_section_info(ws, row, stripped[5:].strip(), 3)
                i += 1
                continue

            # 表格开始
            if stripped.startswith("|") and i + 1 < len(ch_lines) and re.match(r"^\|[\s\-:|]+\|$", ch_lines[i + 1].strip()):
                # 收集连续表格行
                table_lines = []
                while i < len(ch_lines) and ch_lines[i].strip().startswith("|"):
                    table_lines.append(ch_lines[i])
                    i += 1
                table_rows = parse_table(table_lines)
                if table_rows:
                    table_num += 1
                    table_total += 1
                    row = write_table(ws, row, table_rows, table_num)
                continue

            # 普通段落（非空）
            if stripped and not stripped.startswith(">") and not stripped.startswith("---"):
                cell = ws.cell(row=row, column=1, value=stripped)
                cell.font = NORMAL_FONT
                cell.alignment = LEFT_WRAP
                row += 1
            elif stripped.startswith("---"):
                row += 1

            i += 1

        # 自适应列宽（基于第一行表头长度）
        max_cols = ws.max_column
        for col in range(1, max(max_cols, 1) + 1):
            max_len = 10
            for r in range(1, min(ws.max_row + 1, 200)):
                val = ws.cell(row=r, column=col).value
                if val:
                    # 中文按2计算
                    length = sum(2 if ord(c) > 127 else 1 for c in str(val))
                    if length > max_len:
                        max_len = length
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        # 冻结首行
        ws.freeze_panes = "A2"

    print(f"共写入 {table_total} 个表格")
    print(f"保存到: {DST}")
    wb.save(DST)
    print("完成!")


if __name__ == "__main__":
    main()
